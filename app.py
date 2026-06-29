"""
Zabbix Desktop Monitor — app.py
"""

import sys
import csv
import subprocess
import logging
import logging.handlers
from datetime import datetime

from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QStackedWidget, QMessageBox, QFileDialog, QComboBox,
    QSpinBox, QCheckBox, QTextEdit, QSplitter, QStatusBar,
    QMainWindow, QTabWidget, QHeaderView, QSizePolicy, QFrame,
)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QShortcut, QKeySequence

import matplotlib
matplotlib.use('QtAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas

from zabbix_backend import ZabbixBackend

# ── openpyxl: авто-установка ─────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font as XFont, PatternFill, Alignment
    XLSX_OK = True
except ImportError:
    try:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl
        from openpyxl.styles import Font as XFont, PatternFill, Alignment
        XLSX_OK = True
    except Exception:
        XLSX_OK = False

# ── Тёмная тема ──────────────────────────────────────────────────────────────
DARK = """
QWidget {
    background-color: #13151f;
    color: #dde1f0;
    font-family: 'Segoe UI', 'Arial', sans-serif;
    font-size: 13px;
}
QMainWindow, QStackedWidget { background-color: #13151f; }

QLineEdit, QSpinBox, QTextEdit, QComboBox {
    background: #1e2235; border: 1px solid #2e3352;
    border-radius: 7px; padding: 6px 10px; color: #dde1f0;
    selection-background-color: #4c6ef5;
}
QLineEdit:focus, QSpinBox:focus, QTextEdit:focus, QComboBox:focus {
    border: 1px solid #4c6ef5;
}
QComboBox::drop-down { border: none; width: 24px; }
QComboBox QAbstractItemView {
    background: #1e2235; border: 1px solid #2e3352;
    selection-background-color: #4c6ef5; outline: none;
}
QPushButton {
    background: #1e2235; border: 1px solid #2e3352;
    border-radius: 7px; padding: 7px 16px;
    color: #dde1f0; font-weight: 500;
}
QPushButton:hover  { background: #4c6ef5; border-color: #4c6ef5; color: #fff; }
QPushButton:pressed { background: #3b55d4; }
QPushButton#btnPrimary {
    background: #4c6ef5; border-color: #4c6ef5;
    color: #fff; font-size: 14px; font-weight: 700; padding: 11px 20px;
}
QPushButton#btnPrimary:hover { background: #3b55d4; }
QPushButton#btnPrimary:disabled { background: #252a40; color: #5a6080; border-color: #2e3352; }
QPushButton#btnDanger { background: #c0392b; border-color: #c0392b; color: #fff; }
QPushButton#btnDanger:hover { background: #a93226; }

QTableWidget {
    background: #1a1d2e; alternate-background-color: #1e2235;
    gridline-color: #2e3352; border: 1px solid #2e3352; border-radius: 8px;
}
QTableWidget::item { padding: 5px 8px; border: none; }
QTableWidget::item:selected { background: #2c3566; color: #fff; }
QHeaderView::section {
    background: #252a40; color: #8892c0;
    font-weight: 600; font-size: 11px;
    padding: 6px 8px; border: none;
    border-bottom: 1px solid #2e3352;
}

QTabWidget::pane { border: 1px solid #2e3352; border-radius: 8px; background: #1a1d2e; }
QTabBar::tab {
    background: #1e2235; border: 1px solid #2e3352;
    border-bottom: none; border-radius: 6px 6px 0 0;
    padding: 7px 18px; margin-right: 3px; color: #8892c0;
}
QTabBar::tab:selected { background: #1a1d2e; color: #dde1f0; border-bottom: 2px solid #4c6ef5; }
QTabBar::tab:hover:!selected { background: #252a40; color: #dde1f0; }

QSplitter::handle { background: #2e3352; width: 1px; }

QStatusBar {
    background: #0d0f18; color: #5a6080; font-size: 12px;
    border-top: 1px solid #2e3352; padding: 2px 8px;
}
QStatusBar::item { border: none; }

QCheckBox { spacing: 6px; }
QCheckBox::indicator {
    width: 16px; height: 16px;
    border: 1px solid #2e3352; border-radius: 4px; background: #1e2235;
}
QCheckBox::indicator:checked { background: #4c6ef5; border-color: #4c6ef5; }

QScrollBar:vertical { background: #13151f; width: 8px; margin: 0; }
QScrollBar::handle:vertical {
    background: #2e3352; border-radius: 4px; min-height: 30px;
}
QScrollBar::handle:vertical:hover { background: #4c6ef5; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

QLabel#metaLabel { color: #8892c0; font-size: 11px; }
"""


# ── Потоки ────────────────────────────────────────────────────────────────────
class HostLoaderThread(QThread):
    done = pyqtSignal(list)

    def __init__(self, backend):
        super().__init__()
        self.backend = backend

    def run(self):
        # Полностью в фоне — UI не блокируется
        try:
            hosts = self.backend.get_all_hosts()
        except Exception:
            hosts = []
        self.done.emit(hosts)


class MetricLoaderThread(QThread):
    done = pyqtSignal(str, list, list, str)   # slot_name, times, values, real_name

    def __init__(self, backend, host_id, slot_name, hours):
        super().__init__()
        self.backend   = backend
        self.host_id   = host_id
        self.slot_name = slot_name
        self.hours     = hours

    def run(self):
        try:
            t, v, real_name = self.backend.get_metric_history(
                self.host_id, self.slot_name, self.hours)
        except Exception:
            t, v, real_name = [], [], None
        self.done.emit(self.slot_name, t, v, real_name or '')


class HostTypeThread(QThread):
    """Определяет, есть ли у хоста агент, и возвращает 4 имени слотов метрик."""
    done = pyqtSignal(list)

    def __init__(self, backend, host_id):
        super().__init__()
        self.backend = backend
        self.host_id = host_id

    def run(self):
        try:
            slots = self.backend.get_metric_slots(self.host_id)
        except Exception:
            slots = ['CPU', 'RAM', 'Диск', 'Сеть']
        self.done.emit(slots)


# ── График ────────────────────────────────────────────────────────────────────
class MetricChart(FigureCanvas):
    # Цвета по позиции графика (порядок слотов), а не по фиксированному имени —
    # так как название слота теперь может быть любым (CPU или ICMP ping).
    PALETTE = ['#4c6ef5', '#37b24d', '#f59f00', '#e64980']

    def __init__(self, position=0, parent=None):
        self.position    = position
        self.slot_name   = ''
        self.color       = self.PALETTE[position % len(self.PALETTE)]
        self.fig, self.ax = plt.subplots(figsize=(5, 2.5))
        self.fig.patch.set_facecolor('#1a1d2e')
        super().__init__(self.fig)
        self.setParent(parent)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._empty()

    def set_slot(self, slot_name):
        self.slot_name = slot_name

    def _style(self):
        self.ax.set_facecolor('#13151f')
        self.ax.tick_params(colors='#8892c0', labelsize=8)
        for s in self.ax.spines.values():
            s.set_color('#2e3352')

    def _empty(self, msg=None):
        self.ax.clear()
        label = self.slot_name or '—'
        self.ax.text(0.5, 0.5, msg or f'{label}: выберите узел',
                     ha='center', va='center', fontsize=10, color='#5a6080',
                     transform=self.ax.transAxes)
        self.ax.set_xticks([]); self.ax.set_yticks([])
        self._style()
        self.fig.tight_layout(pad=0.6)
        self.draw()

    @staticmethod
    def _guess_ylabel(slot_name):
        n = slot_name.lower()
        if 'loss' in n:
            return '%'
        if 'response time' in n or 'ping' in n and 'response' not in n:
            return ''
        if 'time' in n:
            return 'мс'
        if 'utilization' in n or '%' in slot_name:
            return '%'
        return ''

    def update_chart(self, times, values, real_name=''):
        self.ax.clear()
        title = real_name or self.slot_name
        if times and values:
            x = list(range(len(times)))
            self.ax.plot(x, values, color=self.color, linewidth=1.8)
            self.ax.fill_between(x, values, alpha=0.12, color=self.color)
            step = max(1, len(times) // 7)
            self.ax.set_xticks(x[::step])
            self.ax.set_xticklabels(times[::step], rotation=25, ha='right')
            self.ax.set_ylabel(self._guess_ylabel(title), color='#8892c0', fontsize=9)
            self.ax.set_title(title, color='#dde1f0', fontsize=10, pad=4)
            self.ax.grid(True, linestyle='--', alpha=0.15, color='#4c6ef5')
            self.ax.set_xlim(0, max(1, len(times) - 1))
        else:
            self._empty(f'{title}: нет данных\n(метрика не найдена в Zabbix)')
            return
        self._style()
        self.fig.tight_layout(pad=0.6)
        self.draw()


# ── Окно входа ────────────────────────────────────────────────────────────────
class LoginWindow(QWidget):
    def __init__(self, backend, on_success):
        super().__init__()
        self.backend    = backend
        self.on_success = on_success
        self._build()

    def _build(self):
        self.setWindowTitle('Zabbix Desktop Monitor')
        self.setFixedSize(440, 380)

        # Главный layout — растягивается и центрирует карточку
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addStretch(1)

        # Горизонтальное центрирование карточки
        h = QHBoxLayout()
        h.addStretch(1)

        # Карточка
        card = QFrame()
        card.setFixedWidth(340)
        card.setStyleSheet('''
            QFrame {
                background: #1e2235;
                border: 1px solid #2e3352;
                border-radius: 14px;
            }
        ''')
        card_v = QVBoxLayout(card)
        card_v.setContentsMargins(32, 32, 32, 32)
        card_v.setSpacing(14)
        card_v.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Иконка
        lbl_icon = QLabel('◈')
        lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_icon.setStyleSheet('font-size: 44px; color: #4c6ef5; background: transparent; border: none;')
        card_v.addWidget(lbl_icon)

        # Заголовок
        lbl_title = QLabel('Zabbix Desktop Monitor')
        lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_title.setStyleSheet(
            'font-size: 17px; font-weight: 700; color: #dde1f0; background: transparent; border: none;')
        card_v.addWidget(lbl_title)

        # Подзаголовок
        lbl_sub = QLabel('Мониторинг сетевой инфраструктуры')
        lbl_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_sub.setStyleSheet(
            'font-size: 11px; color: #5a6080; margin-bottom: 8px; background: transparent; border: none;')
        card_v.addWidget(lbl_sub)

        # Поля
        self.txt_login = QLineEdit()
        self.txt_login.setPlaceholderText('Логин')
        self.txt_login.setText('Admin')
        self.txt_login.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.txt_login.textChanged.connect(self._upd)
        card_v.addWidget(self.txt_login)

        self.txt_pw = QLineEdit()
        self.txt_pw.setPlaceholderText('Пароль')
        self.txt_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self.txt_pw.setText('zabbix')
        self.txt_pw.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.txt_pw.textChanged.connect(self._upd)
        self.txt_pw.returnPressed.connect(self._login)
        card_v.addWidget(self.txt_pw)

        self.btn = QPushButton('Подключиться к API')
        self.btn.setObjectName('btnPrimary')
        self.btn.clicked.connect(self._login)
        card_v.addWidget(self.btn)

        h.addWidget(card)
        h.addStretch(1)
        outer.addLayout(h)
        outer.addStretch(1)

        self._upd()

    def _upd(self):
        ok = bool(self.txt_login.text()) and bool(self.txt_pw.text())
        self.btn.setEnabled(ok)

    def _login(self):
        self.btn.setText('Подключение…')
        self.btn.setEnabled(False)
        QApplication.processEvents()
        if self.backend.connect(self.txt_login.text().strip(), self.txt_pw.text()):
            self.on_success()
        else:
            QMessageBox.critical(self, 'Ошибка',
                'Не удалось подключиться к Zabbix API!\n'
                'Проверьте логин, пароль и сетевое соединение.')
        self.btn.setText('Подключиться к API')
        self._upd()


# ── Главное окно ──────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self, backend):
        super().__init__()
        self.backend    = backend
        self.all_hosts  = []
        self._host_id   = None
        self._threads   = []   # держим ссылки чтобы GC не убрал
        self._charts    = []   # 4 виджета MetricChart, заполняются в _build()
        self._is_loading_hosts = False   # защита от повторного запуска load_hosts
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._filter)
        self._build()
        self._shortcuts()

    def _build(self):
        self.setWindowTitle('Zabbix Desktop Monitor')
        self.resize(1280, 780)

        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(10)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        root.addWidget(splitter)

        # ── Левая панель ─────────────────────────────────────────────────
        lp = QWidget()
        lp.setMinimumWidth(340)
        lp.setMaximumWidth(500)
        lv = QVBoxLayout(lp)
        lv.setContentsMargins(0, 0, 0, 0)
        lv.setSpacing(7)

        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText('🔍  Поиск по имени или IP…')
        self.txt_search.textChanged.connect(self._on_search_changed)
        lv.addWidget(self.txt_search)

        self.cmb_group = QComboBox()
        self.cmb_group.addItem('— Все группы —')
        self.cmb_group.currentTextChanged.connect(self._filter)
        lv.addWidget(self.cmb_group)

        br = QHBoxLayout()
        self.btn_refresh = QPushButton('⟳  Обновить')
        self.btn_refresh.clicked.connect(self.load_hosts)
        br.addWidget(self.btn_refresh)
        self.btn_csv = QPushButton('📄 CSV')
        self.btn_csv.clicked.connect(self._export_csv)
        br.addWidget(self.btn_csv)
        self.btn_xlsx = QPushButton('📊 Excel')
        self.btn_xlsx.clicked.connect(self._export_xlsx)
        br.addWidget(self.btn_xlsx)
        lv.addLayout(br)

        self.tbl = QTableWidget()
        self.tbl.setColumnCount(4)
        self.tbl.setHorizontalHeaderLabels(['Имя узла', 'IP-адрес', 'Группа', 'Состояние'])
        self.tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.verticalHeader().setVisible(False)
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.currentCellChanged.connect(self._host_selected)
        lv.addWidget(self.tbl)

        ar = QHBoxLayout()
        self.chk_auto = QCheckBox('Авто каждые')
        self.chk_auto.stateChanged.connect(self._toggle_auto)
        ar.addWidget(self.chk_auto)
        self.spn = QSpinBox()
        self.spn.setRange(5, 3600)
        self.spn.setValue(30)
        self.spn.setSuffix(' сек')
        self.spn.setFixedWidth(90)
        self.spn.valueChanged.connect(self._update_auto)
        ar.addWidget(self.spn)
        ar.addStretch()
        lv.addLayout(ar)

        splitter.addWidget(lp)

        # ── Правая панель ─────────────────────────────────────────────────
        rp = QWidget()
        rv = QVBoxLayout(rp)
        rv.setContentsMargins(0, 0, 0, 0)
        rv.setSpacing(7)

        self.lbl_title = QLabel('Выберите узел из списка слева')
        self.lbl_title.setStyleSheet(
            'font-size: 15px; font-weight: 700; color: #4c6ef5; padding: 4px 0;')
        rv.addWidget(self.lbl_title)

        self.tabs = QTabWidget()
        rv.addWidget(self.tabs)

        # Вкладка Метрики
        mw = QWidget()
        mv = QVBoxLayout(mw)
        mv.setContentsMargins(8, 8, 8, 8)
        mv.setSpacing(7)

        ir = QHBoxLayout()
        ir.addWidget(QLabel('Интервал:'))
        self.cmb_int = QComboBox()
        self.cmb_int.addItems(['1 час', '3 часа', '6 часов', '12 часов', '24 часа'])
        self.cmb_int.setFixedWidth(130)
        self.cmb_int.currentIndexChanged.connect(self._refresh_charts)
        ir.addWidget(self.cmb_int)
        btn_rd = QPushButton('Обновить графики')
        btn_rd.clicked.connect(self._refresh_charts)
        ir.addWidget(btn_rd)
        ir.addStretch()
        mv.addLayout(ir)

        grid = QHBoxLayout()
        grid.setSpacing(6)
        lc = QVBoxLayout()
        rc = QVBoxLayout()

        self.ch_0 = MetricChart(0)
        self.ch_1 = MetricChart(1)
        self.ch_2 = MetricChart(2)
        self.ch_3 = MetricChart(3)

        lc.addWidget(self.ch_0)
        lc.addWidget(self.ch_1)
        rc.addWidget(self.ch_2)
        rc.addWidget(self.ch_3)
        self._charts = [self.ch_0, self.ch_1, self.ch_2, self.ch_3]
        grid.addLayout(lc)
        grid.addLayout(rc)
        mv.addLayout(grid)
        self.tabs.addTab(mw, '📊  Метрики')

        # Вкладка Комментарии
        cw = QWidget()
        cv = QVBoxLayout(cw)
        cv.setContentsMargins(12, 12, 12, 12)
        cv.setSpacing(8)
        cv.addWidget(QLabel('Комментарий к узлу:'))
        self.txt_comment = QTextEdit()
        self.txt_comment.setPlaceholderText('Заявка №…, причина, контакт…')
        self.txt_comment.setMaximumHeight(150)
        cv.addWidget(self.txt_comment)
        cbr = QHBoxLayout()
        bs = QPushButton('💾  Сохранить')
        bs.setObjectName('btnPrimary')
        bs.clicked.connect(self._save_comment)
        cbr.addWidget(bs)
        bd = QPushButton('🗑  Удалить')
        bd.setObjectName('btnDanger')
        bd.clicked.connect(self._del_comment)
        cbr.addWidget(bd)
        cbr.addStretch()
        cv.addLayout(cbr)
        cv.addStretch()
        self.tabs.addTab(cw, '💬  Комментарии')

        splitter.addWidget(rp)
        splitter.setSizes([380, 900])

        self.sb = QStatusBar()
        self.setStatusBar(self.sb)
        self.sb.showMessage('Готово к работе')

        self._auto_timer = QTimer()
        self._auto_timer.timeout.connect(self.load_hosts)

    def _shortcuts(self):
        QShortcut(QKeySequence('Ctrl+F'), self, activated=lambda: self.txt_search.setFocus())
        QShortcut(QKeySequence('Ctrl+E'), self, activated=self._export_csv)
        QShortcut(QKeySequence('Ctrl+R'), self, activated=self.load_hosts)

    # ── Загрузка хостов (неблокирующая) ─────────────────────────────────────
    def load_hosts(self):
        if self._is_loading_hosts:
            return   # предыдущая загрузка ещё не завершилась — пропускаем тик
        self._is_loading_hosts = True
        self.sb.showMessage('⏳  Загрузка списка узлов…')
        self.btn_refresh.setEnabled(False)
        t = HostLoaderThread(self.backend)
        t.done.connect(self._hosts_loaded)
        t.start()
        self._threads.append(t)

    def _hosts_loaded(self, hosts):
        self.all_hosts = hosts

        cur = self.cmb_group.currentText()
        self.cmb_group.blockSignals(True)
        self.cmb_group.clear()
        self.cmb_group.addItem('— Все группы —')
        groups = sorted({h['group'] for h in hosts if h['group'] not in ('—', '')})
        self.cmb_group.addItems(groups)
        idx = self.cmb_group.findText(cur)
        self.cmb_group.setCurrentIndex(idx if idx >= 0 else 0)
        self.cmb_group.blockSignals(False)

        self._filter()
        now = datetime.now().strftime('%H:%M:%S')
        self.sb.showMessage(
            f'✅  Обновлено в {now}  —  узлов: {len(hosts)}')
        self.btn_refresh.setEnabled(True)
        self._is_loading_hosts = False

        # Очищаем завершённые потоки
        self._threads = [t for t in self._threads if t.isRunning()]

    # ── Фильтрация ───────────────────────────────────────────────────────────
    def _on_search_changed(self):
        # Debounce: перезапускаем таймер на каждое нажатие,
        # реальная фильтрация выполнится через 250 мс простоя
        self._search_timer.start(250)

    def _filter(self):
        text  = self.txt_search.text().lower().strip()
        group = self.cmb_group.currentText()
        all_g = (group == '— Все группы —')
        filtered = [
            h for h in self.all_hosts
            if (text in h['name'].lower() or text in h['ip'])
            and (all_g or h['group'] == group)
        ]
        self._display(filtered)

    def _display(self, hosts):
        self.tbl.blockSignals(True)
        try:
            self.tbl.setRowCount(len(hosts))
            for row, h in enumerate(hosts):
                if h['has_problems']:
                    bg = QColor('#2d1a1a')
                    st, sc = '⚠ Проблема', QColor('#e03131')
                elif h['available']:
                    bg = QColor('#1a2a1e')
                    st, sc = '● Доступен', QColor('#37b24d')
                else:
                    bg = QColor('#1e2235')
                    st, sc = '○ Выключен', QColor('#868e96')

                for col, val in enumerate([h['name'], h['ip'], h['group'], st]):
                    item = QTableWidgetItem(val)
                    item.setBackground(bg)
                    if col == 0:
                        item.setData(Qt.ItemDataRole.UserRole, h['id'])
                    if col == 3:
                        item.setForeground(sc)
                    self.tbl.setItem(row, col, item)
            self.tbl.resizeRowsToContents()
        finally:
            self.tbl.blockSignals(False)

    # ── Выбор хоста ──────────────────────────────────────────────────────────
    def _host_selected(self, row, *_):
        item = self.tbl.item(row, 0)
        if not item:
            return
        self._host_id = item.data(Qt.ItemDataRole.UserRole)
        self.lbl_title.setText(f'Узел: {item.text()}')
        self._refresh_charts()
        self.txt_comment.setPlainText(self.backend.get_comment(self._host_id))

    def _get_hours(self):
        return {'1 час': 1, '3 часа': 3, '6 часов': 6,
                '12 часов': 12, '24 часа': 24}.get(self.cmb_int.currentText(), 1)

    def _refresh_charts(self):
        if not self._host_id:
            return
        for chart in self._charts:
            chart.setVisible(True)
            chart._empty('Определение типа узла…')

        self.sb.showMessage('⏳  Определение доступных метрик узла…')
        host_id = self._host_id   # фиксируем — пока поток бежит, выбор могли сменить
        t = HostTypeThread(self.backend, host_id)
        t.done.connect(lambda slots, hid=host_id: self._slots_ready(hid, slots))
        t.start()
        self._threads.append(t)

    def _slots_ready(self, host_id, slots):
        # Если пользователь уже выбрал другой узел — отбрасываем устаревший результат
        if host_id != self._host_id:
            return
        hours = self._get_hours()

        if not slots:
            for chart in self._charts:
                chart.set_slot('')
                chart._empty('Метрики не найдены\n(у узла нет данных в Zabbix)')
            self.sb.showMessage('⚠  У узла нет ни одной известной метрики')
            return

        self.sb.showMessage(f'⏳  Загрузка метрик ({len(slots)} найдено)…')

        # Заполняем графики по числу найденных слотов; лишние — очищаем
        for i, chart in enumerate(self._charts):
            if i < len(slots):
                slot_name = slots[i]
                chart.set_slot(slot_name)
                chart.setVisible(True)
                chart._empty(f'{slot_name}: загрузка…')
                t = MetricLoaderThread(self.backend, host_id, slot_name, hours)
                t.done.connect(lambda sn, tm, vl, rn, hid=host_id: self._metric_done(hid, sn, tm, vl, rn))
                t.start()
                self._threads.append(t)
            else:
                chart.set_slot('')
                chart._empty('—')
                chart.setVisible(False)

    def _metric_done(self, host_id, slot_name, times, values, real_name):
        if host_id != self._host_id:
            return   # узел сменился — игнорируем устаревший результат
        for chart in self._charts:
            if chart.slot_name == slot_name:
                chart.update_chart(times, values, real_name)
                break
        self.sb.showMessage('✅  Метрики загружены')
        self._threads = [t for t in self._threads if t.isRunning()]

    # ── Экспорт ──────────────────────────────────────────────────────────────
    def _visible_hosts(self):
        result = []
        for row in range(self.tbl.rowCount()):
            item = self.tbl.item(row, 0)
            if item:
                hid = item.data(Qt.ItemDataRole.UserRole)
                match = next((h for h in self.all_hosts if h['id'] == hid), None)
                if match:
                    result.append(match)
        return result

    def _export_csv(self):
        if not self.all_hosts:
            return
        path, _ = QFileDialog.getSaveFileName(self, 'Сохранить CSV', '', 'CSV (*.csv)')
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(['ID', 'Имя', 'IP', 'Группа', 'Статус'])
                for h in self._visible_hosts():
                    w.writerow([h['id'], h['name'], h['ip'], h['group'],
                                 'Активен' if h['available'] else 'Выключен'])
            self.sb.showMessage(f'✅  CSV сохранён: {path}')
        except PermissionError:
            QMessageBox.critical(self, 'Ошибка',
                'Нет прав на запись в указанную папку.')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    def _export_xlsx(self):
        if not XLSX_OK:
            QMessageBox.critical(self, 'Ошибка',
                'Запустите: pip install openpyxl')
            return
        if not self.all_hosts:
            return
        path, _ = QFileDialog.getSaveFileName(self, 'Сохранить Excel', '', 'Excel (*.xlsx)')
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Хосты Zabbix'
            hf = PatternFill('solid', fgColor='3B4FBF')
            hfont = XFont(bold=True, color='FFFFFF')
            for col, name in enumerate(['ID', 'Имя', 'IP', 'Группа', 'Статус'], 1):
                c = ws.cell(row=1, column=col, value=name)
                c.font = hfont; c.fill = hf
                c.alignment = Alignment(horizontal='center')
            pf  = PatternFill('solid', fgColor='3D1A1A')
            of  = PatternFill('solid', fgColor='1A2D1E')
            dif = PatternFill('solid', fgColor='1E2235')
            for row, h in enumerate(self._visible_hosts(), 2):
                ws.append([h['id'], h['name'], h['ip'], h['group'],
                           'Активен' if h['available'] else 'Выключен'])
                fill = pf if h['has_problems'] else (of if h['available'] else dif)
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fill
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = (
                    max(len(str(c.value or '')) for c in col) + 4)
            wb.save(path)
            self.sb.showMessage(f'✅  Excel сохранён: {path}')
        except PermissionError:
            QMessageBox.critical(self, 'Ошибка', 'Нет прав на запись в папку.')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', str(e))

    # ── Комментарии ──────────────────────────────────────────────────────────
    def _save_comment(self):
        if not self._host_id:
            return
        self.backend.save_comment(self._host_id, self.txt_comment.toPlainText())
        self.sb.showMessage('✅  Комментарий сохранён')

    def _del_comment(self):
        if not self._host_id:
            return
        self.backend.delete_comment(self._host_id)
        self.txt_comment.clear()
        self.sb.showMessage('🗑  Комментарий удалён')

    # ── Автообновление ───────────────────────────────────────────────────────
    def _toggle_auto(self, state):
        if state == Qt.CheckState.Checked.value:
            self._auto_timer.start(self.spn.value() * 1000)
            self.sb.showMessage(f'⟳  Автообновление каждые {self.spn.value()} сек')
        else:
            self._auto_timer.stop()
            self.sb.showMessage('⏹  Автообновление остановлено')

    def _update_auto(self, val):
        if self._auto_timer.isActive():
            self._auto_timer.start(val * 1000)


# ── Контроллер ────────────────────────────────────────────────────────────────
class Controller:
    def __init__(self):
        self.backend = ZabbixBackend(server_url='http://192.168.101.220/zabbix')
        self.stack   = QStackedWidget()
        self.stack.setWindowTitle('Zabbix Desktop Monitor')
        self.stack.resize(440, 380)

        self.login = LoginWindow(self.backend, self._show_main)
        self.main  = MainWindow(self.backend)

        self.stack.addWidget(self.login)
        self.stack.addWidget(self.main)
        self.stack.setCurrentWidget(self.login)
        self.stack.show()

    def _show_main(self):
        screen = QApplication.primaryScreen().geometry()
        self.stack.resize(1280, 780)
        self.stack.move((screen.width() - 1280) // 2,
                        (screen.height() - 780) // 2)
        self.stack.setCurrentWidget(self.main)
        self.main.load_hosts()


# ── Точка входа ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    plt.rcParams.update({
        'text.color': '#dde1f0', 'axes.labelcolor': '#8892c0',
        'xtick.color': '#8892c0', 'ytick.color': '#8892c0',
        'axes.edgecolor': '#2e3352', 'figure.facecolor': '#1a1d2e',
        'axes.facecolor': '#13151f',
    })
    app = QApplication(sys.argv)
    app.setStyleSheet(DARK)
    controller = Controller()
    sys.exit(app.exec())
